# This file has all the function definitions used by FEMMAPP.PY
# IMPORT some libraries
from openpyxl import load_workbook
from scipy.interpolate import LinearNDInterpolator
from cmath import rect, exp, polar
from math import pi, sqrt
from matplotlib.colors import ListedColormap
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import cm

#----------------------------------------------------------------------------------------------------------------------
# Function for loading data from excel file


def femmdata(sheet_name, xl_fp):

    wb = load_workbook(xl_fp, data_only=True)
    ws = wb[sheet_name]

    d1 = []; d2 = []; d3 = []; d4 = []

    # Read main details from Excel (3rd row)
    for x in range(1, 9): d1.append(ws.cell(row=3, column=x).value)

    numt = d1[5]    # Number of Terminals
    numl = d1[6]    # Number of Layers

    for x in range(1, 4):                   # Read Terminal details, rows = 6, 7, 8
        temp = []
        for y in range(1, numt+1):
            temp.append(ws.cell(row=x+5, column=y+1).value)
        d2.append(temp)

    for x in range(1, 8):                   # Read Layer details, rows = 11 to 17
        temp = []
        for y in range(1, numl+1):
            temp.append(ws.cell(row=x+10, column=y+1).value)
        d3.append(temp)

    max_seg = np.max(d3[3])      # Maximum number of segments

    for x in range(1, 9):                   # Read Segment details, rows = 20 to 27
        temp = []
        for y in range(1, max_seg+1):
            temp.append(ws.cell(row=x+19, column=y+1).value)
        d4.append(temp)

    return d1, d2, d3, d4

#----------------------------------------------------------------------------------------------------------------------

# Calculate Flux Density from Vector potential(Phi = 2.pi.r.A)..
# Code: Praviraj PG, Rev 0, Date: June-11-2022..
# Reference: "FemmviewDoc.cpp", Function: "GetElementB", FEMM Scource Code..
# A = Real Component of Vector Potential
# Fr = Interpolated Grid Values, Radial component of Flux
# Fz = Interpolated Grid Values, Axial component of Flux
# Fa = Interpolated Grid Values, Vector Potential
# E = Total Energy Stored in air


def femmflux():

    in2m = 0.0254       # Inches to Meters..

    ansf = 'femmxfr.ans'   # Read output file from FEMM - this is the *.ANS file created by FEMM.
    with open(ansf, 'r') as file:
        raw = file.readlines()      # Raw data from FEMM Answer file

    R = [];   Z = [];   A = []; Ar = []
    n1 = [];    n2 = [];    n3 = []
    ln = 0; kk = 0; nn = 0
    flag = False
    for line in raw:
        ln += 1

        if '[Solution]' in line:
            flag = True; kk = ln
            continue

        if flag and ln == kk+1:
            nn = int(line)         # Number of Nodes
            continue

        if flag and kk + 1 < ln <= kk + nn + 1:
            res = line.split()
            R.append(float(res[0])); Z.append(float(res[1]))
            A.append(float(res[2]) + float(res[3])*1j)
            Ar.append(float(res[2]))
            # A = Ar + 1i*Ai;         # Vector Potential is COMPLEX for AC Problems (Freq > 0)
            continue

        if flag and ln == kk+nn+2:
            ne = int(line)  # Number of Elements
            continue

        if flag and kk + nn + 2 < ln <= kk + nn + ne + 2:
            res = line.split()
            n1.append(int(res[0])); n2.append(int(res[1])); n3.append(int(res[2]))

    # Calculate Flux Density..
    E = 0      # Stored Energy
    Br = []; Bz = []     # Store Radial(Br) & Axial (Bz) of an element..
    Rr = []; Zz = []     # Store R & Z, coordinates of centroid of triangle element..

    for mm in range(ne):

        # Get R, Z, A values of nodes of the element..
        r1 = R[n1[mm]]; r2 = R[n2[mm]]; r3 = R[n3[mm]]
        z1 = Z[n1[mm]]; z2 = Z[n2[mm]]; z3 = Z[n3[mm]]
        a1 = A[n1[mm]]; a2 = A[n2[mm]]; a3 = A[n3[mm]]

        # Compute Area..
        b1 = z2-z3;     b2 = z3-z1;     b3 = z1-z2
        c1 = r3-r2;     c2 = r1-r3;     c3 = r2-r1
        da = (b1*c2 - b2*c1)
        r = (r1+r2+r3)/3        # Element Centroid, r
        z = (z1+z2+z3)/3        # Element Centroid, z

        # Construct values for mid-side nodes
        if r1<1e-6 and r2<1e-6:
            v1 = (a1+a2)/2
        else:
            v1 = (r2*(3*a1+a2) + r1*(a1+3*a2))/(4*(r1+r2))

        if r2<1e-6 and r3<1e-6:
            v2 = (a2+a3)/2
        else:
            v2 = (r3*(3*a2+a3) + r2*(a2+3*a3))/(4*(r2+r3))

        if r3<1e-6 and r1<1e-6:
            v3 = (a3+a1)/2
        else:
            v3 = (r1*(3*a3+a1) + r3*(a3+3*a1))/(4*(r3+r1))

        # derivatives w.r.t. p and q
        dp = (-a1 + a2 + 4*v2 - 4*v3)/3
        dq = (-a1 - 4*v1 + 4*v2 + a3)/3

        # Compute Flux for Each element..
        da = da * 2 * pi * r * in2m**2
        tBr = -(c2*dp + c3*dq)/da
        tBz = (b2*dp + b3*dq)/da

        # Compute Energy Stored E = 1/2 x (B2/mu) x Vol, exclude Core
        E = E + (tBr**2 + tBz**2) * da * in2m / (4 * pi * 4e-7)

        Br.append(tBr); Bz.append(tBz)
        Rr.append(r);  Zz.append(z)

    Fr = LinearNDInterpolator(list(zip(Rr, Zz)), Br)    # Interpolated Grid Values, Radial component of Flux
    Fz = LinearNDInterpolator(list(zip(Rr, Zz)), Bz)    # Interpolated Grid Values, Axial component of Flux
    Fa = LinearNDInterpolator(list(zip(R, Z)), Ar)    # Interpolated Grid Values, Vector Potential

    return E, Rr, Zz, Fr, Fz, Fa

#----------------------------------------------------------------------------------------------------------------------

# Instantaneous Eddy & Flux Density for each segment,..
# Code: Praviraj PG, Jun-11-2022


def femmeddy(d1, d3, d4, Js, Fr, Fz):

    units = d1[0]; xx = 1                # Units
    if units == 2: xx = 25.4        # Use xx as conversion factor for mm to inches..
    f = d1[2]                       # Frequency..
    numl = d1[6]                    # No. of Layers..
    ms = np.max(d3[3])         # Total segments..
    cr = 0.0008				        # Corner radius reduction, sq.in

    #---------------------------------------------------------------
    # Formula used:
    # Average Eddy (see eq 4.102, 4.103 from Transformer Book by S.V.Kulkarni, in SI units)
    # Per = (2*pi*f*Br*cw)^2/(24*rh); # Eddy per cycle per unit volume due to radial flux in Watts/m3
    # Pea = (2*pi*f*Bz*ct)^2/(24*rh); # Eddy per cycle unit volume due to axial flux in Watts/m3
    # Since frequency & resistivity are same for every segment, formula is simplified as below:
    #  Per = ex *(Br*cw)^2; Pea = ex *(Bz*ct)^2;
    #  where ex = 0.685*f^2/rh;   Also Converted to US units
    #---------------------------------------------------------------

    # Each segments..
    Pe = np.zeros((ms, 2))               # Eddy loss..
    I2R = np.zeros((ms, 1))				# I2R Loss...
    Ns = np.zeros((ms, 4))				# Strands per turn..
    sn = 1;                         # Starts with first segment..
    for ly in range(numl):

        ir = d3[1][ly] / 2 / xx         # Inside Radius..
        rd = d3[2][ly] / xx             # Radial..
        ls = d3[3][ly]                  # Last Segment..
        tp = d3[6][ly]				    # Conductor Type.. 1 = Copper, 2 = Aluminum
        rh = 0.825 if tp == 1 else 1.353
        # Resistivity of Copper (0.825) vs Aluminum (1.353) at 75°C in miroOhm-inch
        ex = 0.685 * f**2 / rh		# Constants for eddy loss estimation..

        # Calculation for Each segment..
        while sn <= ls:

            # Segments
            x1 = ir;            x2 = ir + rd
            y1 = d4[0][sn - 1] / xx;            y2 = d4[1][sn - 1] / xx
            nt = d4[2][sn-1]       			# Number of Turns in the segment..
            at = d4[3][sn-1]       			# Number of Active Turns in the segment..
            st = d4[4][sn-1]       			# Strands/Turn..
            rl = d4[5][sn-1]       			# Radial Strands/Layer..
            ct = d4[6][sn-1]/xx;    			# Conductor Thickness..
            cw = d4[7][sn-1]/xx;    			# Conductor Width..
            mv = 2 * pi * (ct*cw-cr)     	# mv * radial loc. = Volume per turn = turn length x area..
            rc = rh * 2 * pi / (ct*cw-cr)*1e-6	    # Resistance per turn per strand...

            T1 = rd/rl             # Divide radial into # of strands radially..
            # Divide axial height into approx. axial strands/discs..
            T2 = np.around(nt*st/rl);  T2 = (y2-y1)/T2
            mm = np.arange(x1+T1/2, x2, T1)      # Divide points radially - locations of conductors radially..
            nn = np.arange(y1+T2/2, y2, T2)      # Divide points axially - locations of condcutors axially..
            m = len(mm);             # Number of points radially..
            n = len(nn);             # Number of points axially..
            # The segment is divided into a [m x n] matrix where Flux densities will be extracted..
            Pea = 0;    Per = 0;  Ir = 0
            for p in range(m):
                for q in range(n):

                    #Btemp = mo_getb(mm[p],nn[q])   # Get Br & Bz values directly from FEMM, may be slower
                    #Br = Btemp[0]             # Get Radial flux density..
                    #Bz = Btemp[1]             # Get Axial flux density..

                    # or Get Values from femmflux.FLUXBRZ function
                    Br = Fr(mm[p], nn[q])
                    Bz = Fz(mm[p], nn[q])

                    # Eddy loss
                    Per = Per + ex *(Br*cw)**2 * mv * mm[p]   # Eddy - radial flux..
                    Pea = Pea + ex *(Bz*ct)**2 * mv * mm[p]   # Eddy - axial flux..
                    Ir = Ir + (Js[sn-1]/st/sqrt(2))**2 * mm[p]*rc			# I2R Loss per strand..

            #Ns(sn,:) = [st rl ct cw];		# Strands per turn..
            ft = (nt*st)/(m*n)     # A multiplier corresponding to actual number of conductor locations..
            Pe[sn-1][0] = abs(Pea) * ft;    Pe[sn-1][1] = abs(Per) * ft      # Each Segment Eddy x multiplier..
            I2R[sn-1][0] = Ir * ft * at/nt			# I2R Loss for Active turns..
            sn = sn + 1        # go for next segment..

    return Pe, I2R

#----------------------------------------------------------------------------------------------------------------------

# FEMM Analysis Report..
# Code: Praviraj PG, Jun-11-2022


def femmreport(sheet_name,d1,d2,d3,d4,AT,X,E,I2R,Pe,a_r):

    units = d1[0]       # Units..
    MVA = d1[1]         # MVA
    ff = d1[2]           # Frequency
    Di = d1[3]
    Z = d1[4]
    numt = d1[5]        # No. of Terminals..
    numl = d1[6]        # No. of Layers..
    dt = d1[7]          # Distance to tank..

    unit = 'INCH' if units == 1 else 'MM'

    # Initialize a string for storing details
    rep = ""
    # MVA & Details
    rep += f'#################################################################################\n'
    rep += f'                  FEM ANALYSIS ({ff:2} Hz) REPORT - JOB # {sheet_name:5} \n'
    rep += f'---------------------------------------------------------------------------------\n'
    rep += f'                                   INPUT \n'
    rep += f'---------------------------------------------------------------------------------\n'
    rep += f'| INPUT |  BASE  |  FREQ. |   CORE   |  WINDOW  |      NUMBER OF     | DISTANCE |\n'
    rep += f'| UNITS |   MVA  |   HZ   |   DIA.   |  HEIGHT  | TERMINALS | LAYERS |  TO TANK |\n'
    rep += f'  {unit:4}    {MVA:7.3f} \t{ff:3}     {Di:8.3f}   {Z:8.3f} \t  {numt:2} \t\t{numl:2} \t\t{dt:7.2f} \n\n'

    # Terminal Details..
    rep += f'| TERMINAL # |  CONNECTION |    MVA    |     KV    | I COIL ANGLE(°) |\n'
    for g in range(numt):
        contype = d2[0][g]

        if contype == 1: con = 'WYE'                    # 1 - Wye
        elif contype == 2: con = 'DELTA'                # 2 - Delta
        elif contype == 3: con = 'AUTO-HV'              # 3 - Auto High Voltage Terminal
        elif contype == 4: con = 'AUTO-LV'              # 4 - Auto Low Voltage Terminal
        elif contype == 5: con = 'ZIG'                  # 5 - Zig, Neutral Connected Winding
        elif contype == 6: con = 'ZAG'                  # 6 - Zag, Terminal Connected Winding
        elif contype == 7: con = 'POLY.D MAIN'          # 7 - Polygon Delta Main/Inside Winding
        elif contype == 8: con = 'POLY.D TERM'          # 8 - Polygon Delta Terminal Connected Winding
        elif contype == 9: con = 'EXT.D MAIN'           # 9 - Extended Delta Main/Inside Winding
        elif contype == 10: con = 'EXT.D TERM'          # 10 - Extended Delta Terminal Connected Winding
        else : con = 'N/A'

        rep += f' \t {g+1:2} \t   {con:11}   {d2[1][g]:8.3f}     {d2[2][g]:8.3f} \t\t {d2[3][g]:7.2f} \n'

    rep += f'\n'

        # Layer Details..
    rep += f'| LAYER | LAST |  INNER  |  RADIAL | TERM. | PARALL | CURR.  | CONDUCTOR |\n'
    rep += f'|   #   | SEG. |   DIA.  |   WIDTH |   #   | GROUP  | DIRECT |    TYPE   |\n'
    for g in range(numl):
        tp = d3[6][g]
        cond = 'CU' if tp == 1 else 'AL'
        rep += f'   {g+1:2} \t  {d3[3][g]:2} \t {d3[1][g]:7.2f}   {d3[2][g]:7.3f}' \
               f'   {d3[0][g]:2} \t  {d3[4][g]:2} \t   {d3[5][g]:2} \t \t   {cond:2} \n'

    rep += f'\n'

    # Segment Data..
    ms = np.max(d3[3])    # Total segments
    rep += f'| SEG.| LAY |   Z-COORDINATE   |  NO. OF TURNS  |  # STRANDS/   |  STRAND DIM.  |\n'
    rep += f'|  #  |  #  |  ZMIN  |   ZMAX  | TOTAL | ACTIVE | TURN | RADIAL | THICK | WIDTH |\n'
    for g in range(ms):
        rep += f'  {g+1:2} \t{AT[g][1]:2.0f}  {d4[0][g]:7.2f}   {d4[1][g]:7.2f}   {d4[2][g]:6.1f}' \
               f'  {d4[3][g]:6.1f}    {d4[4][g]:2}  \t   {d4[5][g]:2}  \t   {d4[6][g]:5.3f}  {d4[7][g]:6.3f} \n'

    # Output
    # Ampere Turns & Volts/Turn Calculation..
    tAT = []
    for x in range(len(AT)): tAT.append(rect(AT[x][2], AT[x][3] * pi / 180))
    atp = [x for x in tAT if x.real > 0]  # Total Positive Ampere Turns
    atp = sum(atp)
    dev = abs(sum(tAT) / atp) * 100     # Ampere Turn Deviation (%)
    atp = abs(atp)                      # Ampere Turns
    vpt = a_r*MVA * 1e6 / 3 / atp        # V/T = VI/NI, where a_r = auto-ratio
    Ed = 3 * sum(sum(Pe))               # Total Eddy..
    DC = 3 * sum(sum(I2R))              # Total I2R Loss...
    rep += f'---------------------------------------------------------------------------------\n\n'
    rep += f'                               OUTPUT @ {ff:2} Hz \n'
    rep += f'-----------------------------------------------------------------------------\n'
    rep += f'|   3P   |  1P AT |  VOLTS/ | % REACT.(X) | 1P MAGNETIC  | LOSSES (W) @75°C |\n'
    rep += f'|   MVA  |  N X I |   TURN  |    1/2LI2   | ENERGY (J-S) | 3P EDDY | 3P I2R |\n'
    rep += f'  {MVA:6.2f}  {atp:7.0f}   {vpt:7.3f}     {abs(X):6.2f}%    \t  {abs(E):7.1f}  \t ' \
           f'  {Ed:6.0f}   {DC:7.0f}\n\n'

    # Eddy
    rep += f'| SEG.|    1P    |  AMPERE | ANGLE | EDDY LOSS (W) DUE TO  |   TOTAL   |  I2R  |\n'
    rep += f'|  #  |   KVA    |  TURNS  |  (°)  | AXI. FLUX | RAD. FLUX |  EDDY (W) |  (W)  |\n'
    for g in range(ms):
        rep += f'  {g+1:2}    {(AT[g][2]*vpt/1000):8.1f}   {AT[g][2]:7.0f}  {AT[g][3]:7.2f}   {Pe[g][0]:8.2f} '\
               f'    {Pe[g][1]:8.2f}    {(Pe[g][0]+Pe[g][1]):7.1f} {I2R[g][0]:7.0f}\n'

    rep += f'--------------------------------------------------------------------------------\n'
    rep += f'    AMPERE TURNS DEVIATION = {dev:7.3f}% \n'
    rep += f'################################################################################\n'

    with open('report.txt', "w") as file:
        file.write(rep)

#----------------------------------------------------------------------------------------------------------------------

# Custom Colormap - Praviraj PG
# Removes BLUE color


def colmap():

    st = 0.05
    x1 = np.arange(0, 1, st)
    x2 = np.arange(1, 0, -st)

    N = len(x1)
    vals1 = np.ones((N,4)); vals2 = np.ones((N,4))
    vals1[:, 0] = x1;   vals1[:, 2] = x2
    vals2[:, 1] = x2;   vals2[:, 2] = x1
    Y = np.vstack((vals1, vals2))

    return Y

#----------------------------------------------------------------------------------------------------------------------

# Plot Flux Density & Vector potential..
# Code: Praviraj PG, Rev 0, Date: June-11-2022


def femmplot(sheet_name, Fr, Fz, Fa, Ri, R, Z, Ly, xx):

    # Plot Flux Density, Vector Potential & Segments..
    tr = np.linspace(Ri, R, 150)
    tz = np.linspace(0, Z, 200)
    [rr, zz] = np.meshgrid(tr, tz)

    # get flux densities & Vector Potential at Locs - xx,yy
    Br = Fr(rr, zz);     Bz = Fz(rr, zz);         A = abs(Fa(rr, zz))
    B = abs(np.sqrt(np.square(Br) + np.square(Bz)))              # Resultant Flux, T

    # Mesh Plot of Flux Density
    #plt.style.use('ggplot')
    cmap = ListedColormap(colmap())
    #cmap = cm.Greys     # Ref: https://matplotlib.org/3.5.0/tutorials/colors/colormaps.html
    fig = plt.figure(figsize=(4, 9), layout='tight');    ax = fig.gca()
    ax.pcolor(rr*xx, zz*xx, B, cmap=cmap, linewidth=0)
    m = cm.ScalarMappable(cmap=cmap);    m.set_array(B)
    plt.colorbar(m, aspect=30, pad=0.1, label='Flux Density (T)')

    # Equi-potential lines plot
    ax.contour(rr*xx, zz*xx, A, levels=17, linewidths=0.5, colors='k')

    # Plot winding Segments...
    ms = len(Ly)
    for mm in range(ms):
        x1 = Ly[mm][0]*xx; y1 = Ly[mm][1]*xx; x2 = Ly[mm][2]*xx; y2 = Ly[mm][3]*xx
        xc = [x1, x1, x2, x2, x1]; yc = [y1, y2, y2, y1, y1]
        J = Ly[mm][4]; color_seg = 'blue' if J == 0 else 'red'
        ax.plot(xc, yc, linewidth=1.25, zorder=10, color=color_seg)

    # Vector Quiver plot
    tr = np.linspace(Ri, R, 10);    tz = np.linspace(0, Z, 20)
    [rr, zz] = np.meshgrid(tr, tz)
    Br = Fr(rr, zz);    Bz = Fz(rr, zz)
    ax.quiver(rr*xx, zz*xx, Br, Bz, width=0.007, headwidth=7)

    # Some labels
    units = "(inches)" if xx == 1 else "(mm)"
    x_label = f'Radial {units}' + r'$\rightarrow$'
    y_label = f'Core : Window Height {units}' + r'$\rightarrow$'
    ax.set_xlabel(x_label); ax.set_ylabel(y_label)
    plt.title(f'Flux Density Plot : {sheet_name}')
    plt.savefig("femmplot.png")
    #plt.show()

    return plt

#----------------------------------------------------------------------------------------------------------------------